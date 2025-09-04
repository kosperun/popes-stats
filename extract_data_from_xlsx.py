from openpyxl import load_workbook

import csv

wb = load_workbook("Who were the Popes.xlsx")
sheet = wb.active


excel_data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if (row[3] and row[3].lower()) == "interregnum" or (row[5] and row[5].lower() == "interregnum"):
        continue
    if row[1] == 1378:
        excel_data.append(("", "", ""))  # Personal name, Place of birth
    name = f'{row[7].replace(";", ",")}' if row[7] else ""
    excel_data.append((name, row[8], row[9]))  # Personal name, Place of birth

print("length of rows", len(excel_data))

# Read the CSV and add the new columns to each row
with open("popes_dataset.csv", "r") as csv_file:
    csv_reader = list(csv.reader(csv_file))
    header = csv_reader[0] + ["Personal Name", "Place of Birth", "Modern-day country of birth"]
    csv_rows = csv_reader[1:]
    print("length of csv", len(csv_rows))

    # Append Excel data to CSV rows
    for i, row in enumerate(csv_rows):
        row += list(excel_data[i])

# Write back the modified data to the CSV file
with open("popes_dataset.csv", "w", newline="") as csv_file:
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(header)
    csv_writer.writerows(csv_rows)
